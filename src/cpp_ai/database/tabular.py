"""Generic CSV and Excel importers.

These handle arbitrary user datasets given a :class:`ColumnMapping`. CSV uses
the standard library (streaming, no heavy dependency); Excel uses ``openpyxl``
(installed via the ``database`` extra).
"""

from __future__ import annotations

import csv
from pathlib import Path
from typing import Any, Iterator, Mapping

from .base import IMPORTER_REGISTRY, ColumnMapping, Importer


class CSVImporter(Importer):
    """Import peptides from a delimited text file.

    The delimiter is sniffed from the first kilobyte, falling back to a comma,
    so tab- and comma-separated exports both work without configuration.
    """

    dataset_name = "csv"

    def __init__(
        self,
        column_mapping: ColumnMapping | None = None,
        *,
        dataset_name: str | None = None,
        delimiter: str | None = None,
    ) -> None:
        super().__init__(column_mapping)
        if dataset_name is not None:
            self.dataset_name = dataset_name
        self._delimiter = delimiter

    def _sniff_delimiter(self, path: Path) -> str:
        if self._delimiter is not None:
            return self._delimiter
        with path.open("r", newline="", encoding="utf-8-sig") as handle:
            sample = handle.read(1024)
        if not sample:
            return ","
        try:
            return csv.Sniffer().sniff(sample, delimiters=",\t;|").delimiter
        except csv.Error:
            return ","

    def _iter_raw_rows(self, path: Path) -> Iterator[Mapping[str, Any]]:
        delimiter = self._sniff_delimiter(path)
        with path.open("r", newline="", encoding="utf-8-sig") as handle:
            reader = csv.DictReader(handle, delimiter=delimiter)
            for row in reader:
                yield dict(row)


class ExcelImporter(Importer):
    """Import peptides from the first (or a named) worksheet of an .xlsx file."""

    dataset_name = "excel"

    def __init__(
        self,
        column_mapping: ColumnMapping | None = None,
        *,
        dataset_name: str | None = None,
        sheet: str | None = None,
    ) -> None:
        super().__init__(column_mapping)
        if dataset_name is not None:
            self.dataset_name = dataset_name
        self._sheet = sheet

    def _iter_raw_rows(self, path: Path) -> Iterator[Mapping[str, Any]]:
        try:
            from openpyxl import load_workbook
        except ImportError as exc:  # pragma: no cover - dependency guard
            raise ImportError(
                "ExcelImporter requires openpyxl. Install with: "
                'pip install "cpp-ai[database]"'
            ) from exc

        workbook = load_workbook(path, read_only=True, data_only=True)
        try:
            worksheet = workbook[self._sheet] if self._sheet else workbook.active
            rows = worksheet.iter_rows(values_only=True)
            try:
                header = next(rows)
            except StopIteration:
                return
            columns = [str(h) if h is not None else f"col_{i}" for i, h in enumerate(header)]
            for values in rows:
                yield {columns[i]: values[i] if i < len(values) else None for i in range(len(columns))}
        finally:
            workbook.close()


IMPORTER_REGISTRY.register("csv", CSVImporter)
IMPORTER_REGISTRY.register("excel", ExcelImporter)
